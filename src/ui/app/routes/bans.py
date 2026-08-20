from collections import Counter, defaultdict
from contextlib import suppress
from datetime import datetime
from io import BytesIO, StringIO
from ipaddress import ip_address as validate_ip_address
from json import JSONDecodeError, loads
from math import floor
from time import time
from traceback import format_exc
from html import escape, unescape

from flask import Blueprint, Response, jsonify, redirect, render_template, request, send_file, url_for
from flask_login import login_required

from app.api_client import ApiClientError, ApiUnavailableError
from app.dependencies import API_CLIENT, BW_CONFIG, BW_INSTANCES_UTILS
from app.utils import LOGGER, RESERVED_SERVICE_NAMES, csv_safe, csv_writer, flash

from app.routes.utils import (
    cors_required,
    get_default_ban_time,
    get_remain,
    handle_error,
    parse_search_panes,
    parse_search_panes_dict,
)

bans = Blueprint("bans", __name__)


# Column order shared between the table and exports — must stay in sync with bans.js
_BAN_COLUMNS = (
    "date",  # 0
    "ip",  # 1
    "country",  # 2
    "reason",  # 3
    "scope",  # 4
    "service",  # 5
    "end_date",  # 6
    "time_left",  # 7
    "actions",  # 8
)


def _collect_all_bans():
    """Every ban the operator has decided on, enriched with `remain`/`start_date`/`end_date` for
    display. Returns the raw (unfiltered) list.

    Read from the database through the API. It used to scan Redis and fan out to the instances,
    dedup the two, and hope one of them still had the ban: a restarted instance enumerates an empty
    shared dict, and under Redis it only re-materializes a ban when a request from that IP arrives.
    """
    try:
        bans_list = API_CLIENT.get_bans()
    except (ApiClientError, ApiUnavailableError) as e:
        LOGGER.debug(format_exc())
        LOGGER.error(f"Couldn't get bans from the API: {e}")
        return []

    timestamp_now = time()

    for ban in bans_list:
        exp = ban.pop("exp", 0)
        if exp == 0 or ban.get("permanent", False):
            ban["remain"] = "permanent"
            ban["permanent"] = True
            ban["end_date"] = "permanent"
        else:
            remain = ("unknown", "unknown") if exp <= 0 else get_remain(exp)
            ban["remain"] = remain[0]
            ban["start_date"] = datetime.fromtimestamp(floor(ban["date"])).astimezone().isoformat()
            ban["end_date"] = datetime.fromtimestamp(floor(timestamp_now + exp)).astimezone().isoformat()
        # Preserve `exp` for end_date pane filters that still need it
        ban["exp"] = exp

    return bans_list


def _to_float(value, default=0.0):
    try:
        if isinstance(value, (int, float)):
            return float(value)
        if value is None:
            return float(default)
        return float(str(value))
    except Exception:
        return float(default)


def _filter_and_sort_bans(all_bans, search_value, search_panes, order_column_index, order_direction):
    """Apply the same filtering + sorting logic as `/bans/fetch`, returning the
    full filtered list (no pagination)."""

    def filter_by_search_panes(items):
        filtered = items
        for field, selected_values in search_panes.items():
            if not selected_values:
                continue
            if field == "date":
                now = time()

                def date_filter(ban):
                    ban_date = ban.get("date", 0)
                    for val in selected_values:
                        if val == "last_24h" and now - ban_date < 86400:
                            return True
                        if val == "last_7d" and now - ban_date < 604800:
                            return True
                        if val == "last_30d" and now - ban_date < 2592000:
                            return True
                        if val == "older_30d" and now - ban_date >= 2592000:
                            return True
                    return False

                filtered = list(filter(date_filter, filtered))
            elif field == "scope":

                def scope_filter(ban):
                    return ban.get("ban_scope") in selected_values

                filtered = list(filter(scope_filter, filtered))
            elif field == "end_date":

                def end_date_filter(ban):
                    if ban.get("permanent", False) and "future_30d" in selected_values:
                        return True
                    exp = ban.get("exp", 0)
                    if ban.get("permanent", False):
                        return "permanent" in selected_values
                    for val in selected_values:
                        if val == "permanent" and ban.get("permanent", False):
                            return True
                        if val == "next_24h" and exp < 86400:
                            return True
                        if val == "next_7d" and exp < 604800:
                            return True
                        if val == "next_30d" and exp < 2592000:
                            return True
                        if val == "future_30d" and exp >= 2592000:
                            return True
                    return False

                filtered = list(filter(end_date_filter, filtered))
            elif field == "service":

                def service_filter(ban):
                    ban_service = ban.get("service")
                    if ban.get("ban_scope") == "global" or ban_service in (None, ""):
                        ban_service = "_"
                    return str(ban_service) in selected_values

                filtered = list(filter(service_filter, filtered))
            else:
                filtered = [b for b in filtered if str(b.get(field, "N/A")) in selected_values]
        return filtered

    def global_search_filter(ban):
        if search_value == "permanent" and ban.get("permanent", False):
            return True
        return any(search_value in str(ban.get(col, "")).lower() for col in _BAN_COLUMNS)

    filtered_bans = list(filter(global_search_filter, all_bans)) if search_value else list(all_bans)
    filtered_bans = list(filter_by_search_panes(filtered_bans))

    if 0 <= order_column_index < len(_BAN_COLUMNS):
        sort_key = _BAN_COLUMNS[order_column_index]
        if sort_key in ("end_date", "time_left"):
            filtered_bans.sort(
                key=lambda x: ("0" if order_direction == "desc" else "z") if x.get("permanent", False) else x.get(sort_key, ""),
                reverse=(order_direction == "desc"),
            )
        elif sort_key == "date":
            filtered_bans.sort(key=lambda x: _to_float(x.get("date", 0.0), 0.0), reverse=(order_direction == "desc"))
        else:
            filtered_bans.sort(key=lambda x: x.get(sort_key, ""), reverse=(order_direction == "desc"))

    return filtered_bans


def _get_filtered_bans(source):
    return _filter_and_sort_bans(
        _collect_all_bans(),
        source.get("search", "").lower(),
        parse_search_panes_dict(source),
        0,
        "desc",
    )


def _get_filtered_report_bans(source):
    if not BW_INSTANCES_UTILS:
        return []

    try:
        config = BW_CONFIG.get_config(methods=False, with_drafts=True) if BW_CONFIG else {}
    except Exception:
        config = {}

    result = BW_INSTANCES_UTILS.get_reports_query(
        start=0,
        length=-1,
        search=source.get("search", "").lower(),
        order_column="date",
        order_dir="desc",
        search_panes=parse_search_panes(source),
        count_only=False,
        include_pane_counts=False,
    )

    targets = []
    seen = set()
    for report in result.get("data", []):
        ip = report.get("ip")
        server_name = str(report.get("server_name") or "_")
        ban_scope = "global" if server_name == "_" else "service"
        service = "" if ban_scope == "global" else server_name
        key = (ip, ban_scope, service)
        if key in seen:
            continue
        seen.add(key)
        targets.append(
            {
                "ip": ip,
                "reason": report.get("reason") or "ui",
                "ban_scope": ban_scope,
                "service": service,
                "exp": get_default_ban_time(config, server_name),
            }
        )

    return targets


@bans.route("/bans", methods=["GET"])
@login_required
def bans_page():
    # Get list of services for the service dropdown in the UI
    services = BW_CONFIG.get_config(global_only=True, methods=False, with_drafts=True, filtered_settings=("SERVER_NAME",))["SERVER_NAME"]
    if isinstance(services, str):
        services = services.split()

    return render_template("bans.html", services=services)


_BANS_STATS_TOP_REASONS = 5


@bans.route("/bans/stats", methods=["POST"])
@login_required
@cors_required
def bans_stats():
    """Summary tiles + "bans by reason" breakdown for the bans dashboard header."""
    try:
        all_bans = _collect_all_bans()
    except BaseException as e:
        LOGGER.debug(format_exc())
        LOGGER.error(f"Couldn't get bans from redis: {e}")
        all_bans = []

    active = len(all_bans)
    expiring_1h = sum(1 for ban in all_bans if not ban.get("permanent", False) and 0 < ban.get("exp", 0) <= 3600)
    countries = len({ban.get("country") for ban in all_bans if ban.get("country")})
    permanent = sum(1 for ban in all_bans if ban.get("permanent", False))

    reason_counts = Counter(str(ban.get("reason") or "unknown") for ban in all_bans)
    reason_breakdown = [
        {"reason": reason, "count": count, "pct": round(count / active * 100) if active else 0}
        for reason, count in reason_counts.most_common(_BANS_STATS_TOP_REASONS)
    ]

    return jsonify(
        {
            "active": active,
            "expiring_1h": expiring_1h,
            "countries": countries,
            "permanent": permanent,
            "reason_breakdown": reason_breakdown,
        }
    )


@bans.route("/bans/fetch", methods=["POST"])
@login_required
@cors_required
def bans_fetch():
    try:
        bans = _collect_all_bans()
    except BaseException as e:
        LOGGER.debug(format_exc())
        LOGGER.error(f"Couldn't get bans from redis: {e}")
        flash("Failed to fetch bans from Redis, see logs for more information.", "error")
        bans = []

    # DataTables parameters
    draw = int(request.form.get("draw", 1))
    start = max(0, int(request.form.get("start", 0)))
    length = max(1, min(int(request.form.get("length", 10)), 1000))
    search_value = request.form.get("search[value]", "").lower()
    # DataTables includes two leading non-data columns (details-control and select)
    # Adjust incoming index to align with backend data columns
    try:
        order_column_index_dt = int(request.form.get("order[0][column]", 0))
    except Exception:
        order_column_index_dt = 0
    order_column_index = max(order_column_index_dt - 2, 0)
    order_direction = request.form.get("order[0][dir]", "desc")
    search_panes = parse_search_panes_dict(request.form)

    # Local alias kept for the formatter / pane-counts code below
    columns = list(_BAN_COLUMNS)

    # Helper: format a ban for DataTable row
    def format_ban(ban):
        # Defensive: some bans may lack some fields
        return {
            "date": datetime.fromtimestamp(floor(ban.get("date", 0))).isoformat() if ban.get("date") else "N/A",
            "ip": escape(str(ban.get("ip", "N/A"))),
            "country": escape(str(ban.get("country", "N/A"))),
            "reason": escape(str(ban.get("reason", "N/A"))),
            "scope": escape(str(ban.get("ban_scope", "global"))),
            "service": escape(str(ban.get("service") or "_")),
            "end_date": "permanent" if ban.get("permanent", False) else escape(str(ban.get("end_date", "N/A"))),
            "time_left": "permanent" if ban.get("permanent", False) else escape(str(ban.get("remain", "N/A"))),
            "permanent": bool(ban.get("permanent", False)),
            "actions": "",  # Actions column for buttons
        }

    filtered_bans = _filter_and_sort_bans(bans, search_value, search_panes, order_column_index, order_direction)

    paginated_bans = filtered_bans if length == -1 else filtered_bans[start : start + length]  # noqa: E203

    # Format for DataTable
    formatted_bans = [format_ban(ban) for ban in paginated_bans]

    # Calculate pane counts (for SearchPanes)
    pane_counts = defaultdict(lambda: defaultdict(lambda: {"total": 0, "count": 0}))

    # Use IP+scope+service as unique ID for bans
    def ban_id(ban):
        service = ban.get("service")
        # Normalize service to "_" for global bans or when service is None
        if ban.get("ban_scope") == "global" or service is None:
            service = "_"
        return f"{ban.get('ip','')}|{ban.get('ban_scope','')}|{service}"  # noqa: E231

    filtered_ids = {ban_id(ban) for ban in filtered_bans}
    for ban in bans:
        for field in columns[1:]:  # skip date
            value = ban.get(field, "N/A")
            # Special handling for service field to normalize global bans
            if field == "service":
                if ban.get("ban_scope") == "global" or value in (None, ""):
                    value = "_"
            if isinstance(value, (dict, list)):
                value = str(value)
            pane_counts[field][value]["total"] += 1
            if ban_id(ban) in filtered_ids:
                pane_counts[field][value]["count"] += 1

    # Prepare SearchPanes options (special formatting for date, country, scope, service, and end_date)
    base_flags_url = url_for("static", filename="img/flags")
    search_panes_options = {}

    # Special handling for date searchpane options
    search_panes_options["date"] = [
        {
            "label": '<span data-i18n="searchpane.last_24h">Last 24 hours</span>',
            "value": "last_24h",
            "total": sum(1 for ban in bans if time() - ban.get("date", 0) < 86400),
            "count": sum(1 for ban in filtered_bans if time() - ban.get("date", 0) < 86400),
        },
        {
            "label": '<span data-i18n="searchpane.last_7d">Last 7 days</span>',
            "value": "last_7d",
            "total": sum(1 for ban in bans if time() - ban.get("date", 0) < 604800),
            "count": sum(1 for ban in filtered_bans if time() - ban.get("date", 0) < 604800),
        },
        {
            "label": '<span data-i18n="searchpane.last_30d">Last 30 days</span>',
            "value": "last_30d",
            "total": sum(1 for ban in bans if time() - ban.get("date", 0) < 2592000),
            "count": sum(1 for ban in filtered_bans if time() - ban.get("date", 0) < 2592000),
        },
        {
            "label": '<span data-i18n="searchpane.older_30d">More than 30 days</span>',
            "value": "older_30d",
            "total": sum(1 for ban in bans if time() - ban.get("date", 0) >= 2592000),
            "count": sum(1 for ban in filtered_bans if time() - ban.get("date", 0) >= 2592000),
        },
    ]

    # Special handling for country searchpane options
    search_panes_options["country"] = []
    for code, counts in pane_counts["country"].items():
        str_code = str(code)
        country_code = str_code.lower()
        is_unknown = str_code in ("unknown", "local", "n/a")
        flag_code = "zz" if is_unknown else country_code
        # Show both the alpha-2 code and the translated country name so users can search by either
        code_text = "N/A" if is_unknown else str_code.upper()
        i18n_key = "not_applicable" if str_code in ("unknown", "local") else str_code.upper()
        fallback_name = "N/A" if is_unknown else str_code
        search_panes_options["country"].append(
            {
                "label": f'<img src="{base_flags_url}/{flag_code}.svg" class="border border-1 p-0 me-1" height="17" />&nbsp;－&nbsp;<span class="me-1"><code>{code_text}</code></span><span data-i18n="country.{i18n_key}">{fallback_name}</span>',
                "value": str_code,
                "total": counts["total"],
                "count": counts["count"],
            }
        )

    # Special handling for scope searchpane options
    search_panes_options["scope"] = [
        {
            "label": '<i class="bx bx-xs bx-globe"></i> <span data-i18n="scope.global">Global</span>',
            "value": "global",
            "total": sum(1 for ban in bans if ban.get("ban_scope") == "global"),
            "count": sum(1 for ban in filtered_bans if ban.get("ban_scope") == "global"),
        },
        {
            "label": '<i class="bx bx-xs bx-server"></i> <span data-i18n="scope.service_specific">Service</span>',
            "value": "service",
            "total": sum(1 for ban in bans if ban.get("ban_scope") == "service"),
            "count": sum(1 for ban in filtered_bans if ban.get("ban_scope") == "service"),
        },
    ]

    # Special handling for service searchpane options
    search_panes_options["service"] = []
    for name, counts in pane_counts["service"].items():
        display_name = "default server" if (not name or name == "_") else escape(str(name))
        search_panes_options["service"].append(
            {
                "label": display_name,
                "value": escape(str(name)),
                "total": counts["total"],
                "count": counts["count"],
            }
        )

    # Special handling for end_date searchpane options
    search_panes_options["end_date"] = [
        {
            "label": '<span data-i18n="searchpane.permanent">Permanent</span>',
            "value": "permanent",
            "total": sum(1 for ban in bans if ban.get("permanent", False) or ban.get("exp", 0) == 0),
            "count": sum(1 for ban in filtered_bans if ban.get("permanent", False) or ban.get("exp", 0) == 0),
        },
        {
            "label": '<span data-i18n="searchpane.next_24h">Next 24 hours</span>',
            "value": "next_24h",
            "total": sum(1 for ban in bans if not ban.get("permanent", False) and ban.get("exp", 0) < 86400),
            "count": sum(1 for ban in filtered_bans if not ban.get("permanent", False) and ban.get("exp", 0) < 86400),
        },
        {
            "label": '<span data-i18n="searchpane.next_7d">Next 7 days</span>',
            "value": "next_7d",
            "total": sum(1 for ban in bans if not ban.get("permanent", False) and ban.get("exp", 0) < 604800),
            "count": sum(1 for ban in filtered_bans if not ban.get("permanent", False) and ban.get("exp", 0) < 604800),
        },
        {
            "label": '<span data-i18n="searchpane.next_30d">Next 30 days</span>',
            "value": "next_30d",
            "total": sum(1 for ban in bans if not ban.get("permanent", False) and ban.get("exp", 0) < 2592000),
            "count": sum(1 for ban in filtered_bans if not ban.get("permanent", False) and ban.get("exp", 0) < 2592000),
        },
        {
            "label": '<span data-i18n="searchpane.future_30d">More than 30 days</span>',
            "value": "future_30d",
            "total": sum(1 for ban in bans if not ban.get("permanent", False) and ban.get("exp", 0) >= 2592000),
            "count": sum(1 for ban in filtered_bans if not ban.get("permanent", False) and ban.get("exp", 0) >= 2592000),
        },
    ]

    # Add any remaining fields from pane_counts
    for field, values in pane_counts.items():
        if field not in search_panes_options:
            search_panes_options[field] = [
                {
                    "label": escape(str(value)),
                    "value": escape(str(value)),
                    "total": counts["total"],
                    "count": counts["count"],
                }
                for value, counts in values.items()
            ]

    # Response
    return jsonify(
        {
            "draw": draw,
            "recordsTotal": len(bans),
            "recordsFiltered": len(filtered_bans),
            "data": formatted_bans,
            "searchPanes": {"options": search_panes_options},
        }
    )


def _bans_export_rows():
    """Collect, filter, and sort bans using the request's search/searchPanes/order
    parameters and return them as plain dict rows ready to be written to CSV/XLSX."""
    bans = _collect_all_bans()

    search_value = request.args.get("search", "").lower()
    order_column = request.args.get("order_column", "date").strip().lower()
    order_dir = request.args.get("order_dir", "desc").strip().lower()
    if order_dir not in ("asc", "desc"):
        order_dir = "desc"
    try:
        order_column_index = _BAN_COLUMNS.index(order_column)
    except ValueError:
        order_column_index = 0

    search_panes = parse_search_panes_dict(request.args)
    filtered_bans = _filter_and_sort_bans(bans, search_value, search_panes, order_column_index, order_dir)

    rows = []
    for ban in filtered_bans:
        date_value = "N/A"
        ban_date = ban.get("date")
        if ban_date:
            with suppress(Exception):
                date_value = datetime.fromtimestamp(floor(ban_date)).astimezone().isoformat()

        ban_scope = ban.get("ban_scope") or "global"
        service = ban.get("service") or "_"
        if ban_scope == "global" or service in (None, ""):
            service = "_"

        if ban.get("permanent", False):
            end_date = "permanent"
            time_left = "permanent"
        else:
            end_date = unescape(str(ban.get("end_date", "N/A")))
            time_left = unescape(str(ban.get("remain", "N/A")))

        rows.append(
            {
                "date": date_value,
                "ip": str(ban.get("ip", "N/A")),
                "country": str(ban.get("country", "N/A")),
                "reason": str(ban.get("reason", "N/A")),
                "scope": ban_scope,
                "service": service,
                "end_date": end_date,
                "time_left": time_left,
            }
        )
    return rows


_BAN_EXPORT_HEADERS = (
    "Date",
    "IP Address",
    "Country",
    "Reason",
    "Scope",
    "Service",
    "End Date",
    "Time Left",
)
_BAN_EXPORT_FIELDS = ("date", "ip", "country", "reason", "scope", "service", "end_date", "time_left")


@bans.route("/bans/export/csv", methods=["GET"])
@login_required
def bans_export_csv():
    """Export the current filtered+sorted ban list as CSV (all columns, all rows)."""
    try:
        rows = _bans_export_rows()
    except Exception as e:
        LOGGER.error(f"Error collecting bans for CSV export: {e}")
        LOGGER.debug(format_exc())
        return jsonify({"error": "Failed to export bans"}), 500

    output = StringIO()
    writer = csv_writer(output)
    writer.writerow(_BAN_EXPORT_HEADERS)
    for row in rows:
        writer.writerow([row[field] for field in _BAN_EXPORT_FIELDS])

    output.seek(0)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=bunkerweb_bans_{timestamp}.csv"},
    )


@bans.route("/bans/export/excel", methods=["GET"])
@login_required
def bans_export_excel():
    """Export the current filtered+sorted ban list as XLSX (all columns, all rows)."""
    try:
        rows = _bans_export_rows()
    except Exception as e:
        LOGGER.error(f"Error collecting bans for Excel export: {e}")
        LOGGER.debug(format_exc())
        return jsonify({"error": "Failed to export bans"}), 500

    # openpyxl is imported here, not at module scope: it pulls 311 modules and ~120 ms of import
    # time (measured in the UI image, 2026-08-19; 110-210 ms across runs), and this blueprint is
    # registered at startup, so at module scope every UI worker pays that boot cost for an export
    # almost nobody runs.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Bans"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    ws.append(list(_BAN_EXPORT_HEADERS))
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        ws.append([csv_safe(row[field]) for field in _BAN_EXPORT_FIELDS])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            with suppress(Exception):
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"bunkerweb_bans_{timestamp}.xlsx",
    )


def _api_ban(ip: str, exp, reason: str, service: str, ban_scope: str) -> str:
    """Ban through the API, which persists the ban before fanning it out to the instances.

    Returns "" on success or the error to surface. The UI never talks to the instances directly:
    the durable row is what survives a restart, and it is also what lets a later convergence pass
    re-push a ban to an instance that was unreachable now.
    """
    payload = {"ip": ip, "exp": int(exp or 0), "reason": reason}
    if ban_scope == "service" and service:
        payload["service"] = service
    try:
        API_CLIENT.ban([payload])
    except (ApiClientError, ApiUnavailableError) as e:
        return str(e)
    return ""


def _api_unban(ip: str, service, ban_scope: str) -> str:
    """Unban through the API (durable revoke + fan-out). Returns "" on success."""
    payload = {"ip": ip}
    if ban_scope == "service" and service:
        payload["service"] = service
    try:
        API_CLIENT.unban([payload])
    except (ApiClientError, ApiUnavailableError) as e:
        return str(e)
    return ""


@bans.route("/bans/ban", methods=["POST"])
@login_required
def bans_ban():
    # Check database state
    if API_CLIENT.readonly:
        return handle_error("Database is in read-only mode", "bans")

    selection_mode = request.form.get("selection_mode", "explicit")
    if selection_mode == "filtered":
        if request.form.get("source") != "reports":
            return handle_error("Invalid filtered ban source.", "bans", True)
        bans = _get_filtered_report_bans(request.form)
    elif selection_mode == "explicit":
        raw_bans = request.form.get("bans", "")
        if not raw_bans:
            return handle_error("No bans.", "bans", True)
        try:
            bans = loads(raw_bans)
        except JSONDecodeError:
            return handle_error("Invalid bans parameter on /bans/ban.", "bans", True)
    else:
        return handle_error("Invalid ban selection mode.", "bans", True)

    if not bans:
        return handle_error("No matching reports.", "bans", True)

    for ban in bans:
        # Validate ban structure
        if not isinstance(ban, dict) or "ip" not in ban:
            continue

        # Extract and normalize ban parameters
        ip = ban.get("ip", "")
        reason = ban.get("reason", "ui")
        ban_scope = ban.get("ban_scope", "global")
        if ban_scope not in ("global", "service"):  # only two valid scopes; clamp anything else
            ban_scope = "global"
        service = ban.get("service", "")

        # Validate IP address
        try:
            validate_ip_address(ip)
        except ValueError:
            flash(f"Invalid IP address: {escape(ip)}", "error")
            continue

        # Check for permanent ban
        if ban.get("end_date") == "0" or ban.get("exp") == 0:
            ban_end = 0
        else:
            ban_end = ban.get("exp", 0)

        # Validate service name for service-specific bans
        if ban_scope == "service":
            if not service or service in RESERVED_SERVICE_NAMES:
                ban_scope = "global"
                service = "unknown"

        # Persist the ban and propagate it to the connected BunkerWeb instances
        resp = _api_ban(ip, ban_end, reason, service, ban_scope)
        if resp:
            LOGGER.error(f"Failed to ban {ip}: {resp}")
            flash(f"Failed to ban {ip}: {resp}", "error")
        else:
            LOGGER.info(f"Banned {ip}")
            flash(f"Banned {ip} successfully.", "success")

    return redirect(url_for("loading", next=url_for("bans.bans_page"), message=f"Banning {len(bans)} IP{'s' if len(bans) > 1 else ''}"))


@bans.route("/bans/unban", methods=["POST"])
@login_required
def bans_unban():
    # Check database state
    if API_CLIENT.readonly:
        return handle_error("Database is in read-only mode", "bans")

    selection_mode = request.form.get("selection_mode", "explicit")
    if selection_mode == "filtered":
        if request.form.get("source") != "bans":
            return handle_error("Invalid filtered unban source.", "bans", True)
        unbans = [
            {
                "ip": ban.get("ip"),
                "ban_scope": ban.get("ban_scope", "global"),
                "service": ban.get("service"),
            }
            for ban in _get_filtered_bans(request.form)
        ]
    elif selection_mode == "explicit":
        raw_unbans = request.form.get("ips", "")
        if not raw_unbans:
            return handle_error("No bans.", "bans", True)
        try:
            unbans = loads(raw_unbans)
        except JSONDecodeError:
            return handle_error("Invalid ips parameter on /bans/unban.", "bans", True)
    else:
        return handle_error("Invalid unban selection mode.", "bans", True)

    if not unbans:
        return handle_error("No matching bans.", "bans", True)

    for unban in unbans:
        # Validate unban structure
        if "ip" not in unban:
            continue

        # Extract and normalize unban parameters
        ip = unban.get("ip")
        ban_scope = unban.get("ban_scope", "global")
        service = unban.get("service")

        # Validate IP address
        try:
            validate_ip_address(ip)
        except ValueError:
            flash(f"Invalid IP address: {escape(str(ip))}", "error")
            continue

        # Normalize Web UI and default services to global scope
        if service in RESERVED_SERVICE_NAMES:
            ban_scope = "global"
            service = None

        # Revoke the ban durably, then propagate the unban to the instances
        resp = _api_unban(ip, service, ban_scope)
        if resp:
            LOGGER.error(f"Failed to unban {ip}: {resp}")
            flash(f"Failed to unban {ip}: {resp}", "error")
        else:
            LOGGER.info(f"Unbanned {ip}")
            flash(f"Unbanned {ip} successfully.", "success")

    return redirect(url_for("loading", next=url_for("bans.bans_page"), message=f"Unbanning {len(unbans)} IP{'s' if len(unbans) > 1 else ''}"))


@bans.route("/bans/update_duration", methods=["POST"])
@login_required
def bans_update_duration():
    # Check database state
    if API_CLIENT.readonly:
        return handle_error("Database is in read-only mode", "bans")

    selection_mode = request.form.get("selection_mode", "explicit")
    if selection_mode == "filtered":
        if request.form.get("source") != "bans":
            return handle_error("Invalid filtered duration source.", "bans", True)
        duration = request.form.get("duration", "")
        updates = [
            {
                "ip": ban.get("ip"),
                "duration": duration,
                "ban_scope": ban.get("ban_scope", "global"),
                "service": ban.get("service"),
                "custom_exp": request.form.get("custom_exp"),
                "end_date": request.form.get("end_date"),
            }
            for ban in _get_filtered_bans(request.form)
        ]
    elif selection_mode == "explicit":
        raw_updates = request.form.get("updates", "")
        if not raw_updates:
            return handle_error("No updates.", "bans", True)
        try:
            updates = loads(raw_updates)
        except JSONDecodeError:
            return handle_error("Invalid updates parameter on /bans/update_duration.", "bans", True)
    else:
        return handle_error("Invalid duration selection mode.", "bans", True)

    if not updates:
        return handle_error("No matching bans.", "bans", True)

    # Fetch the stored bans to keep each one's original reason across the duration change
    known_bans = {}
    try:
        for ban in API_CLIENT.get_bans():
            ban_key = f"{ban.get('ip')}|{ban.get('ban_scope', 'global')}|{ban.get('service', '_') if ban.get('ban_scope') == 'service' else '_'}"
            known_bans[ban_key] = ban
    except (ApiClientError, ApiUnavailableError) as e:
        LOGGER.error(f"Couldn't read the existing bans, durations will be updated with the default reason: {e}")

    for update in updates:
        # Validate update structure
        if not isinstance(update, dict) or "ip" not in update or "duration" not in update:
            continue

        # Extract and normalize update parameters
        ip = update.get("ip", "")
        duration = update.get("duration", "")
        ban_scope = update.get("ban_scope", "global")
        service = update.get("service", "")

        if duration not in ("permanent", "1h", "24h", "1w", "custom"):
            flash(f"Invalid ban duration: {escape(str(duration))}", "error")
            continue

        # Validate IP address
        try:
            validate_ip_address(ip)
        except ValueError:
            flash(f"Invalid IP address: {escape(ip)}", "error")
            continue

        # Calculate new expiration time based on duration
        if duration == "permanent":
            new_exp = 0
        elif duration == "1h":
            new_exp = 3600
        elif duration == "24h":
            new_exp = 86400
        elif duration == "1w":
            new_exp = 604800
        elif duration == "custom":
            custom_exp = update.get("custom_exp", None)
            if custom_exp is not None:
                try:
                    new_exp = max(0, int(custom_exp))
                except (TypeError, ValueError):
                    flash(f"Invalid custom ban duration for {escape(ip)}", "error")
                    continue
            else:
                custom_end_date = update.get("end_date")
                if custom_end_date:
                    try:
                        end_dt = datetime.fromisoformat(custom_end_date)
                        if end_dt.tzinfo is None:
                            end_dt = end_dt.replace(tzinfo=datetime.now().astimezone().tzinfo)
                        new_exp = max(0, int(end_dt.timestamp() - time()))
                    except (TypeError, ValueError):
                        flash(f"Invalid custom ban end date for {escape(ip)}", "error")
                        continue
                else:
                    flash(f"Missing custom ban end date for {escape(ip)}", "error")
                    continue

        # Validate service name for service-specific bans
        if ban_scope == "service":
            if not service or service in RESERVED_SERVICE_NAMES:
                ban_scope = "global"
                service = "unknown"

        # Fetch existing ban data first to preserve original reason
        original_reason = "ui"  # Default fallback
        ban_key = f"{ip}|{ban_scope}|{service if ban_scope == 'service' else '_'}"
        if ban_key in known_bans:
            original_reason = known_bans[ban_key].get("reason", "ui")

        # Re-ban with the new expiry, keeping the original reason
        ban_resp = _api_ban(ip, new_exp, original_reason, service, ban_scope)
        if ban_resp:
            LOGGER.error(f"Failed to update ban duration for {ip}: {ban_resp}")
            flash(f"Failed to update ban duration for {ip}: {ban_resp}", "error")
        else:
            LOGGER.info(f"Updated ban duration for {ip}")
            flash(f"Updated ban duration for {ip} successfully.", "success")

    return redirect(url_for("loading", next=url_for("bans.bans_page"), message=f"Updating duration for {len(updates)} ban{'s' if len(updates) > 1 else ''}"))
