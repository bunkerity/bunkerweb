from collections import defaultdict
from contextlib import suppress
from datetime import datetime
from json import dumps, loads
from traceback import format_exc
from html import escape
from io import StringIO, BytesIO
from time import monotonic


from flask import Blueprint, jsonify, render_template, request, url_for, Response, send_file
from flask_login import login_required
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.dependencies import API_CLIENT, BW_CONFIG, BW_INSTANCES_UTILS
from app.api_client import ApiClientError, ApiUnavailableError
from app.utils import LOGGER, csv_safe, csv_writer

from app.routes.utils import cors_required, get_default_ban_time, parse_search_panes

_PERSIST_TO_DB_CACHE = {"enabled": True, "expires_at": 0.0}
_PERSIST_TO_DB_TTL_SECONDS = 30.0


def _persist_to_db_enabled() -> bool:
    """Whether reports are persisted to the DB (``METRICS_PERSIST_TO_DB``), cached briefly so we
    don't fetch the config on every reports query."""
    now = monotonic()
    if _PERSIST_TO_DB_CACHE["expires_at"] < now:
        try:
            config = BW_CONFIG.get_config(methods=False) if BW_CONFIG else {}
            _PERSIST_TO_DB_CACHE["enabled"] = config.get("METRICS_PERSIST_TO_DB", "yes") == "yes"
        except Exception:
            _PERSIST_TO_DB_CACHE["enabled"] = True
        _PERSIST_TO_DB_CACHE["expires_at"] = now + _PERSIST_TO_DB_TTL_SECONDS
    return bool(_PERSIST_TO_DB_CACHE["enabled"])


def _legacy_reports(**kwargs):
    """Legacy per-instance/Redis scrape, used when DB persistence is off or the metrics API is down."""
    if BW_INSTANCES_UTILS:
        return BW_INSTANCES_UTILS.get_reports_query(**kwargs)
    return {"total": 0, "filtered": 0, "data": [], "pane_counts": {}}


def _query_reports(**kwargs):
    """Read blocked-request reports.

    With DB persistence enabled (``METRICS_PERSIST_TO_DB``, the default) read from the central API
    (durable DB-backed store), falling back to the legacy scrape only if the API is unavailable. When
    persistence is off the DB isn't populated, so go straight to the legacy per-instance/Redis scrape.
    """
    if not _persist_to_db_enabled():
        return _legacy_reports(**kwargs)
    try:
        return API_CLIENT.get_metrics_requests(**kwargs)
    except (ApiClientError, ApiUnavailableError) as e:
        LOGGER.warning(f"Metrics API unavailable ({e}); falling back to instance scrape")
        return _legacy_reports(**kwargs)


REPORTS_EXPORT_PAGE_SIZE = 10000
REPORTS_EXPORT_MAX_ROWS = 1000000


def _fetch_all_reports(*, search="", order_column="date", order_dir="desc", search_panes=""):
    """Fetch all matching reports for export, paging through ``_query_reports`` so a large export never
    becomes one unbounded, timeout-prone request (which would silently degrade to a truncated scrape)."""
    rows = []
    start = 0
    while len(rows) < REPORTS_EXPORT_MAX_ROWS:
        page = _query_reports(
            start=start,
            length=REPORTS_EXPORT_PAGE_SIZE,
            search=search,
            order_column=order_column,
            order_dir=order_dir,
            search_panes=search_panes,
            count_only=False,
            include_pane_counts=False,
        ).get("data", [])
        rows.extend(page)
        if len(page) < REPORTS_EXPORT_PAGE_SIZE:
            break
        start += REPORTS_EXPORT_PAGE_SIZE
    return rows


reports = Blueprint("reports", __name__)
REPORTS_FILTERS_CACHE_TTL_SECONDS = 5.0
REPORTS_FILTERS_CACHE_MAX_ENTRIES = 128
REPORTS_FILTERS_CACHE: dict[str, dict[str, object]] = {}


def prune_reports_filters_cache(cache_now: float) -> None:
    expired_keys = [key for key, entry in REPORTS_FILTERS_CACHE.items() if float(entry.get("expires_at", 0.0) or 0.0) <= cache_now]
    for key in expired_keys:
        REPORTS_FILTERS_CACHE.pop(key, None)

    if len(REPORTS_FILTERS_CACHE) <= REPORTS_FILTERS_CACHE_MAX_ENTRIES:
        return

    overflow = len(REPORTS_FILTERS_CACHE) - REPORTS_FILTERS_CACHE_MAX_ENTRIES
    oldest_keys = sorted(
        REPORTS_FILTERS_CACHE.items(),
        key=lambda item: float(item[1].get("created_at", 0.0) or 0.0),
    )[:overflow]
    for key, _ in oldest_keys:
        REPORTS_FILTERS_CACHE.pop(key, None)


def build_search_panes_options(pane_counts_backend: dict) -> dict:
    base_flags_url = url_for("static", filename="img/flags")
    search_panes_options = {}

    for field, values in pane_counts_backend.items():
        if field == "country":
            search_panes_options["country"] = []
            for code, counts in values.items():
                str_code = str(code)
                country_code = str_code.lower()
                is_unknown = str_code in ("unknown", "local", "n/a")
                fallback_label = "N/A" if is_unknown else str_code
                i18n_key = "not_applicable" if str_code in ("unknown", "local") else str_code.upper()
                flag_code = "zz" if is_unknown else country_code
                search_panes_options["country"].append(
                    {
                        "label": f'<img src="{base_flags_url}/{flag_code}.svg" class="border border-1 p-0 me-1" height="17" />&nbsp;－&nbsp;<span data-i18n="country.{i18n_key}">{fallback_label}</span>',
                        "value": str_code,
                        "total": counts["total"],
                        "count": counts["count"],
                    }
                )
        elif field == "server_name":
            search_panes_options["server_name"] = []
            for name, counts in values.items():
                search_panes_options["server_name"].append(
                    {
                        "label": f'<code class="language-dns">{escape("default server" if name == "_" else name)}</code>',
                        "value": escape(name),
                        "total": counts["total"],
                        "count": counts["count"],
                    }
                )
        else:
            search_panes_options[field] = [
                {
                    "label": escape(str(value)),
                    "value": escape(str(value)),
                    "total": counts["total"],
                    "count": counts["count"],
                }
                for value, counts in values.items()
            ]

    return search_panes_options


def pane_counts_has_values(pane_counts_backend: dict) -> bool:
    if not isinstance(pane_counts_backend, dict):
        return False
    return any(isinstance(values, dict) and len(values) > 0 for values in pane_counts_backend.values())


@reports.route("/reports", methods=["GET"])
@login_required
def reports_page():
    return render_template("reports.html")


@reports.route("/reports/dashboard", methods=["POST"])
@login_required
@cors_required
def reports_dashboard():
    try:
        end = int(request.form.get("end") or datetime.now().timestamp())
        start = int(request.form.get("start") or (end - 86400))
        bucket = request.form.get("bucket", "hour")
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "Invalid start/end"}), 400

    try:
        timeseries = API_CLIENT.get_metrics_timeseries(start=start, end=end, bucket=bucket)
        offenders = API_CLIENT.get_metrics_top_offenders(start=start, end=end, limit=10)
        rules = API_CLIENT.get_metrics_top_rules(start=start, end=end, limit=10)
    except ApiClientError as e:
        if e.status_code == 400:
            # The API rejected the range itself (e.g. a crafted window needing too many
            # timeseries buckets) -- a client input error, not a service-availability one.
            LOGGER.warning(f"Metrics API rejected the request ({e})")
            return jsonify({"status": "error", "message": str(e) or "Invalid range"}), 400
        LOGGER.warning(f"Metrics API unavailable ({e}); dashboard tabs will show empty state")
        return jsonify({"status": "error", "message": "Metrics service unavailable"}), 503
    except ApiUnavailableError as e:
        LOGGER.warning(f"Metrics API unavailable ({e}); dashboard tabs will show empty state")
        return jsonify({"status": "error", "message": "Metrics service unavailable"}), 503

    return jsonify({"status": "success", "timeseries": timeseries, "offenders": offenders.get("offenders", []), "rules": rules.get("rules", [])})


@reports.route("/reports/fetch", methods=["POST"])
@login_required
@cors_required
def reports_fetch():
    # Load configuration to resolve Bad Behavior ban time per service
    try:
        db_config = BW_CONFIG.get_config(methods=False, with_drafts=True) if BW_CONFIG else {}
    except Exception:
        db_config = {}

    # Extract DataTables parameters
    draw = int(request.form.get("draw", 1))
    start = max(0, int(request.form.get("start", 0)))
    length = max(1, min(int(request.form.get("length", 10)), 1000))
    search_value = request.form.get("search[value]", "").lower()

    # DataTables includes two leading non-data columns (details-control and select)
    # Adjust the incoming order column index to match our data columns mapping
    try:
        order_column_index_dt = int(request.form.get("order[0][column]", 0))
    except Exception:
        order_column_index_dt = 0
    # Subtract 2 to align with backend columns (starting at "date")
    order_column_index = max(order_column_index_dt - 2, 0)
    order_direction = request.form.get("order[0][dir]", "desc")

    # Parse search panes
    search_panes = defaultdict(list)
    for key, value in request.form.items():
        if key.startswith("searchPanes["):
            field = key.split("[")[1].split("]")[0]
            search_panes[field].append(value)

    # Keep this in sync with the frontend DataTables columns
    columns = [
        "date",
        "id",
        "ip",
        "country",
        "method",
        "url",
        "status",
        "user_agent",
        "reason",
        "server_name",
        "data",
        "security_mode",
        "actions",  # actions column for row buttons
    ]

    # Build search panes string for backend (format: field1:value1,value2;field2:value3)
    search_panes_str = ";".join(f"{field}:{','.join(values)}" for field, values in search_panes.items()) if search_panes else ""
    include_pane_counts = False

    # Determine order column name
    order_column_name = columns[order_column_index] if 0 <= order_column_index < len(columns) else "date"

    # Use optimized query endpoint from InstancesUtils
    if BW_INSTANCES_UTILS:
        try:
            result = _query_reports(
                start=start,
                length=length,
                search=search_value,
                order_column=order_column_name,
                order_dir=order_direction,
                search_panes=search_panes_str,
                count_only=False,
                include_pane_counts=include_pane_counts,
            )

            total_count = result.get("total", 0)
            filtered_count = result.get("filtered", 0)
            paginated_reports = result.get("data", [])

        except Exception as e:
            LOGGER.error(f"Error using optimized reports query: {e}")
            LOGGER.debug(format_exc())
            # Fallback to old method
            result = {"total": 0, "filtered": 0, "data": [], "pane_counts": {}}
            total_count = 0
            filtered_count = 0
            paginated_reports = []
    else:
        # Fallback when BW_INSTANCES_UTILS is not available
        result = {"total": 0, "filtered": 0, "data": [], "pane_counts": {}}
        total_count = 0
        filtered_count = 0
        paginated_reports = []

    # Format reports for the response
    def format_report(report):
        try:
            data_field = report.get("data", {})
            has_data = False
            if isinstance(data_field, dict):
                has_data = len(data_field) > 0
            elif isinstance(data_field, list):
                has_data = len(data_field) > 0
            elif isinstance(data_field, str):
                stripped = data_field.strip()
                has_data = stripped not in ("", "{}", "[]", "null", "None")
            else:
                has_data = data_field is not None

            request_id = str(report.get("id", "N/A"))
            server_name = str(report.get("server_name", "N/A"))

            return {
                "date": datetime.fromtimestamp(report.get("date", 0)).isoformat() if report.get("date") else "N/A",
                "id": escape(request_id),
                "request_id": request_id,
                "ip": escape(str(report.get("ip", "N/A"))),
                "country": escape(str(report.get("country", "N/A"))),
                "method": escape(str(report.get("method", "N/A"))),
                "url": escape(str(report.get("url", "N/A"))),
                "status": escape(str(report.get("status", "N/A"))),
                "user_agent": escape(str(report.get("user_agent", "N/A"))),
                "reason": escape(str(report.get("reason", "N/A"))),
                "server_name": escape(server_name),
                # Keep payloads light; details are lazy-loaded by request_id.
                "data": "available" if has_data else "",
                "has_data": has_data,
                "security_mode": escape(str(report.get("security_mode", "N/A"))),
                # default ban duration in seconds for quick-ban action
                "ban_default_exp": get_default_ban_time(db_config, server_name),
                # Placeholder for UI actions column
                "actions": "",
            }
        except Exception as e:
            LOGGER.error(f"Error formatting report: {e}")
            # Return a safe fallback if formatting fails
            return {
                "date": "N/A",
                "id": "N/A",
                "request_id": "N/A",
                "ip": escape(str(report.get("ip", "N/A"))),
                "country": "N/A",
                "method": "N/A",
                "url": "N/A",
                "status": "N/A",
                "user_agent": "N/A",
                "reason": "Error parsing report data",
                "server_name": "N/A",
                "data": "",
                "has_data": False,
                "security_mode": "N/A",
                "ban_default_exp": 86400,
                "actions": "",
            }

    formatted_reports = [format_report(report) for report in paginated_reports]

    # Response
    response_payload = {
        "draw": draw,
        "recordsTotal": total_count,
        "recordsFiltered": filtered_count,
        "data": formatted_reports,
    }

    return jsonify(response_payload)


@reports.route("/reports/filters", methods=["POST"])
@login_required
@cors_required
def reports_filters():
    search_value = request.form.get("search[value]", "").lower()
    search_panes = defaultdict(list)
    for key, value in request.form.items():
        if key.startswith("searchPanes["):
            field = key.split("[")[1].split("]")[0]
            search_panes[field].append(value)

    search_panes_str = ";".join(f"{field}:{','.join(sorted(values))}" for field, values in sorted(search_panes.items())) if search_panes else ""
    has_active_panes = any(values for values in search_panes.values())

    force_refresh = request.form.get("force", "").strip().lower() in ("1", "true", "yes")
    if has_active_panes:
        # Always recompute pane counters while filters are actively selected
        # to keep SearchPanes counts in sync with current selection.
        force_refresh = True

    cache_key = f"search={search_value}|panes={search_panes_str}"
    cache_now = monotonic()
    prune_reports_filters_cache(cache_now)
    cache_entry = REPORTS_FILTERS_CACHE.get(cache_key) or {}
    cache_payload = cache_entry.get("payload")
    cache_expires_at = float(cache_entry.get("expires_at", 0.0) or 0.0)

    if not force_refresh and cache_payload and cache_expires_at > cache_now:
        return jsonify(cache_payload)

    pane_counts_backend = {}
    if BW_INSTANCES_UTILS:
        try:
            result = _query_reports(
                start=0,
                length=0,
                search=search_value,
                order_column="date",
                order_dir="desc",
                search_panes=search_panes_str,
                count_only=True,
                include_pane_counts=True,
            )
            pane_counts_backend = result.get("pane_counts", {})

            # Safety fallback: if count-only pane aggregation unexpectedly returns
            # empty panes, retry through the non-count-only path (historically stable).
            if not pane_counts_has_values(pane_counts_backend):
                fallback_result = _query_reports(
                    start=0,
                    length=1,
                    search=search_value,
                    order_column="date",
                    order_dir="desc",
                    search_panes=search_panes_str,
                    count_only=False,
                    include_pane_counts=True,
                )
                pane_counts_backend = fallback_result.get("pane_counts", {})
        except Exception as e:
            LOGGER.error(f"Error loading reports filters: {e}")
            LOGGER.debug(format_exc())

    payload = {"searchPanes": {"options": build_search_panes_options(pane_counts_backend)}}
    if pane_counts_has_values(pane_counts_backend):
        REPORTS_FILTERS_CACHE[cache_key] = {
            "payload": payload,
            "expires_at": cache_now + REPORTS_FILTERS_CACHE_TTL_SECONDS,
            "created_at": cache_now,
        }
        prune_reports_filters_cache(cache_now)
    else:
        # Don't keep empty pane snapshots around; allow immediate retry.
        REPORTS_FILTERS_CACHE.pop(cache_key, None)
    return jsonify(payload)


@reports.route("/reports/data", methods=["POST"])
@login_required
@cors_required
def report_data_fetch():
    report_id = request.form.get("report_id", "").strip()
    hostname = request.form.get("hostname", "").strip() or None

    if not report_id:
        return jsonify({"status": "error", "message": "Missing report_id"}), 400

    if not BW_INSTANCES_UTILS:
        return jsonify({"status": "error", "message": "Reports service unavailable"}), 503

    try:
        report_data = BW_INSTANCES_UTILS.get_report_data(report_id=report_id, hostname=hostname)
    except Exception as e:
        LOGGER.error(f"Failed to fetch report data for id={report_id}: {e}")
        LOGGER.debug(format_exc())
        return jsonify({"status": "error", "message": "Failed to load report data"}), 500

    if report_data is None:
        return jsonify({"status": "error", "message": "Report data not found"}), 404

    return jsonify({"status": "success", "report_id": report_id, "data": report_data})


@reports.route("/reports/export/csv", methods=["GET"])
@login_required
def reports_export_csv():
    """Export all reports as CSV"""
    try:
        # Get search and order parameters
        search_value = request.args.get("search", "").lower()
        order_column = request.args.get("order_column", "date")
        order_dir = request.args.get("order_dir", "desc")
        search_panes_str = parse_search_panes(request.args)

        # Get all reports (no pagination)
        if BW_INSTANCES_UTILS:
            all_reports = _fetch_all_reports(
                search=search_value,
                order_column=order_column,
                order_dir=order_dir,
                search_panes=search_panes_str,
            )
        else:
            all_reports = []

        # Create CSV in memory
        output = StringIO()
        writer = csv_writer(output)

        # Write header
        writer.writerow(
            [
                "Date",
                "Request ID",
                "IP Address",
                "Country",
                "Method",
                "URL",
                "Status Code",
                "User-Agent",
                "Reason",
                "Server Name",
                "Data",
                "Security Mode",
            ]
        )

        # Write data rows
        for report in all_reports:
            # Format data field
            data_field = report.get("data", {})
            if isinstance(data_field, str):
                try:
                    data_output = dumps(loads(data_field))
                except (ValueError, TypeError):
                    data_output = data_field
            else:
                data_output = dumps(data_field)

            writer.writerow(
                [
                    datetime.fromtimestamp(report.get("date", 0)).isoformat() if report.get("date") else "N/A",
                    str(report.get("id", "N/A")),
                    str(report.get("ip", "N/A")),
                    str(report.get("country", "N/A")),
                    str(report.get("method", "N/A")),
                    str(report.get("url", "N/A")),
                    str(report.get("status", "N/A")),
                    str(report.get("user_agent", "N/A")),
                    str(report.get("reason", "N/A")),
                    str(report.get("server_name", "N/A")),
                    data_output,
                    str(report.get("security_mode", "N/A")),
                ]
            )

        # Prepare response
        output.seek(0)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename=bunkerweb_reports_{timestamp}.csv"},
        )

    except Exception as e:
        LOGGER.error(f"Error exporting reports to CSV: {e}")
        LOGGER.debug(format_exc())
        return jsonify({"error": "Failed to export reports"}), 500


@reports.route("/reports/export/excel", methods=["GET"])
@login_required
def reports_export_excel():
    """Export all reports as Excel"""
    try:
        # Get search and order parameters
        search_value = request.args.get("search", "").lower()
        order_column = request.args.get("order_column", "date")
        order_dir = request.args.get("order_dir", "desc")
        search_panes_str = parse_search_panes(request.args)

        # Get all reports (no pagination)
        if BW_INSTANCES_UTILS:
            all_reports = _fetch_all_reports(
                search=search_value,
                order_column=order_column,
                order_dir=order_dir,
                search_panes=search_panes_str,
            )
        else:
            all_reports = []

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Reports"

        # Style for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        # Write header
        headers = [
            "Date",
            "Request ID",
            "IP Address",
            "Country",
            "Method",
            "URL",
            "Status Code",
            "User-Agent",
            "Reason",
            "Server Name",
            "Data",
            "Security Mode",
        ]
        ws.append(headers)

        # Apply header styling
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for report in all_reports:
            # Format data field
            data_field = report.get("data", {})
            if isinstance(data_field, str):
                try:
                    data_output = dumps(loads(data_field))
                except (ValueError, TypeError):
                    data_output = data_field
            else:
                data_output = dumps(data_field)

            ws.append(
                [
                    csv_safe(datetime.fromtimestamp(report.get("date", 0)).isoformat() if report.get("date") else "N/A"),
                    csv_safe(report.get("id", "N/A")),
                    csv_safe(report.get("ip", "N/A")),
                    csv_safe(report.get("country", "N/A")),
                    csv_safe(report.get("method", "N/A")),
                    csv_safe(report.get("url", "N/A")),
                    csv_safe(report.get("status", "N/A")),
                    csv_safe(report.get("user_agent", "N/A")),
                    csv_safe(report.get("reason", "N/A")),
                    csv_safe(report.get("server_name", "N/A")),
                    csv_safe(data_output),
                    csv_safe(report.get("security_mode", "N/A")),
                ]
            )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                with suppress(Exception):
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"bunkerweb_reports_{timestamp}.xlsx",
        )

    except Exception as e:
        LOGGER.error(f"Error exporting reports to Excel: {e}")
        LOGGER.debug(format_exc())
        return jsonify({"error": "Failed to export reports"}), 500
